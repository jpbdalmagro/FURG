import openpyxl
import json

with open('tabela_conversao.json', 'r', encoding='utf8') as f:
    decodificador = json.load(f)


workbook_matriz = openpyxl.load_workbook('matriz.xlsx')
planilha_matriz = workbook_matriz['Página1']


workbook_nova_matriz = openpyxl.Workbook()
planilha_nova_matriz = workbook_nova_matriz.active
planilha_nova_matriz.title = 'Página1'


for linha in planilha_matriz.iter_rows(min_row=1, max_row=32, min_col=1, max_col=58):
    linha_decodificada = []
    for celula in linha:
        linha_decodificada.append(decodificador[str(int(celula.value))])
        
    planilha_nova_matriz.append(linha_decodificada)

workbook_nova_matriz.save('nova_matriz.xlsx')
print('Matriz decodificada com sucesso!')
