import json
import openpyxl
import pandas as pd


def mensagem():
    mensagem = openpyxl.load_workbook('matriz.xlsx')
    planilha = mensagem.active
    matriz = []
    for linha in planilha.iter_rows(values_only=True):
        matriz.append(linha)
    return matriz


def codigo():
    with open('tabela_conversao.json', 'r', encoding='utf8') as arquivo:
        dict = json.load(arquivo)
    return dict


def chave(caminho):
    with open(caminho, 'r') as arquivo:
        dec = arquivo.read()
        dec = dec.split()
        dec_list = [ [int(dec[n]) for n in range(i, i + 32)] for i in range(0, len(dec), 32) ]
    return dec_list


msg = pd.DataFrame(mensagem())
cripto = codigo()

planilha_msg = openpyxl.load_workbook('matriz_decodificada.xlsx')
planilha = planilha_msg.active

for idx in range(1, 7):
    planilha.append([idx])
    path = f"m{idx}.txt"
    key = chave(path)
    key = pd.DataFrame(key)
    
    decodificado = key.dot(msg)
    decodificado = decodificado.mod(118)
    decodificado = decodificado.applymap(lambda x: cripto[str(x)])
    decodificado = decodificado.T
    decodificado = decodificado.values.tolist()
    
    for linha in decodificado:
        planilha.append(linha)
    
planilha_msg.save('matriz_decodificada.xlsx')
print('Decodificação realizada com sucesso!')
